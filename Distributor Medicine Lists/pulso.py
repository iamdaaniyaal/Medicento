import openpyxl
import re


reader = openpyxl.load_workbook("PULSO ITEM LIST_RAW FORMAT.xlsx")
sheet = reader.active
# arr = list()
print(sheet.max_row)

for i in range(1,sheet.max_row+1):
    try:
        cell_=sheet.cell(i,2)
        arr = cell_.value
        print(arr)
    except:
        continue
